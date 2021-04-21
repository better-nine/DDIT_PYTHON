from openpyxl import load_workbook, Workbook
from flask import Flask,render_template, jsonify, request

app = Flask(__name__, static_url_path="", static_folder="static")

def export_excel():    
    load_wb = load_workbook("score_sheet.xlsx", data_only=True)
    load_ws = load_wb['Sheet1']
    #나온 값이 None일 경우 루프중지
    list =[]
    num = 2
    while load_ws['B'+str(num)].value !=None :
        name = load_ws['A'+str(num)].value
        score = load_ws['B'+str(num)].value
        list.append({'name':name,'score':score})
        num += 1
    print(list)

    return list 

def import_excel(name, score):
    #엑셀에 데이터 삽입
    flag_done = False
    return flag_done

 
write_wb = Workbook()
 
#이름이 있는 시트를 생성
write_ws = write_wb.create_sheet('생성시트')
 
#Sheet1에다 입력
write_ws = write_wb.active
write_ws['A7'] = '테스트'
 
#행 단위로 추가
write_ws.append(['이름테스트','점수테스트'])
 
#셀 단위로 추가
write_ws.cell(1,8,'5행5열')
write_wb.save('score_sheet.xlsx')
 

@app.route('/tetris.html')
def tetris():
    name = request.form["name"]
    score = request.form["score"]
    
    flag = import_excel(name, score) #값 삽입하는 메서드
    list = export_excel() #저장된 데이터 뽑는 메서드
    
    if flag:
        return render_template('tetris.html', list=list)
    else :
        return render_template('tetris.html', err="관리자에게 문의하세요")





