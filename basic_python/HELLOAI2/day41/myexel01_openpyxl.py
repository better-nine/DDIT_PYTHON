from openpyxl import load_workbook, Workbook 

def export_excel():    
    load_wb = load_workbook("myexel.xlsx", data_only=True)
    load_ws = load_wb['시트1']
    #나온 값이 None일 경우 루프중지
    list =[]
    num = 2
    while load_ws['B'+str(num)].value != None :
        name = load_ws['A'+str(num)].value
        number = load_ws['B'+str(num)].value
        list.append({'name':name,'number':number})
        num += 1    
    print(list)
    
    for i, n in enumerate(list):
        print(list[i]['name'],end=" ")
        print(list[i]['number'])
    
    return list 

export_excel()
########################## 이하 선생님 코드 ####################################
load_wb = load_workbook("myexel.xlsx", data_only=True)
load_ws = load_wb['시트1']

print(load_ws['B2'].value)
print("-------------------------------------")
for i in range(2,4):
    print(load_ws.cell(i, 1).value,end=" ")
    print(load_ws.cell(i, 2).value)
    



























